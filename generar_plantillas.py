
"""Script auxiliar para generar las plantillas y archivos de ejemplo."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random

COLUMNAS = [
    "fecha",
    "tipo_comprobante",
    "punto_de_venta",
    "nro_factura",
    "cuit",
    "nombre_empresa",
    "importe_gravado",
    "importe_no_gravado",
    "iva_21",
    "iva_27",
    "iva_105",
    "importe_total",
]

ANCHOS = [14, 18, 14, 16, 20, 30, 16, 18, 10, 10, 12, 16]

COLOR_HEADER = "1A1A2E"
COLOR_EJEMPLO = "E8F0FE"
thin = Side(style="thin", color="CCCCCC")
borde = Border(left=thin, right=thin, top=thin, bottom=thin)


def aplicar_encabezado(ws, titulo, subtitulo):

    """Aplica título y subtítulo a las primeras dos filas de la hoja."""

    ncols = len(COLUMNAS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws["A2"] = subtitulo
    ws["A2"].font = Font(name="Arial", color="888888", size=9, italic=True)
    ws["A2"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16


def aplicar_fila_encabezado(ws, fila=3):

    """Escribe la fila de nombres de columnas con formato."""

    for col, (nombre, ancho) in enumerate(zip(COLUMNAS, ANCHOS), 1):
        c = ws.cell(row=fila, column=col, value=nombre)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="3C3C5A")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borde
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.row_dimensions[fila].height = 20


def escribir_fila_datos(ws, fila, datos, color_bg="FFFFFF"):

    """Escribe una fila de datos con formato."""

    numericas = {"importe_gravado", "importe_no_gravado", "iva_21", "iva_27", "iva_105", "importe_total"}
    for col, (col_name, valor) in enumerate(zip(COLUMNAS, datos), 1):
        c = ws.cell(row=fila, column=col, value=valor)
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor=color_bg)
        c.border = borde
        c.alignment = Alignment(
            horizontal="right" if col_name in numericas else "left",
            vertical="center"
        )
        if col_name in numericas and isinstance(valor, (int, float)):
            c.number_format = '#,##0.00'
    ws.row_dimensions[fila].height = 16


# ── Datos de ejemplo ──────────────────────────────────────────────────────

proveedores = [
    ("20-11111111-1", "ALFA DISTRIBUCIONES SA"),
    ("27-22222222-2", "BETA SERVICIOS SRL"),
    ("30-33333333-3", "GAMMA IMPORTACIONES SA"),
    ("20-44444444-4", "DELTA SUMINISTROS SRL"),
    ("27-55555555-5", "EPSILON TECNOLOGIA SA"),
    ("30-66666666-6", "ZETA LOGISTICA SRL"),
    ("20-77777777-7", "ETA CONSULTORA SA"),
    ("27-88888888-8", "THETA COMERCIAL SRL"),
]

tipos = ["FAC A", "FAC A", "FAC A", "FAC B", "NC A"]


def generar_fila(i, incluir_iva27=False):

    """Genera una fila de datos ficticia para los ejemplos."""

    p = proveedores[i % len(proveedores)]
    tipo = tipos[i % len(tipos)]
    pdv = str((i // 8) + 1).zfill(4)
    nro = str(1000 + i).zfill(8)
    dia = str((i % 28) + 1).zfill(2)
    fecha = f"{dia}/06/2025"

    gravado = round(random.uniform(500, 9000), 2)
    no_gravado = round(random.uniform(0, 300), 2) if i % 4 == 0 else 0.00
    iva_21 = round(gravado * 0.21, 2) if tipo != "FAC B" else 0.00
    iva_27 = round(gravado * 0.27, 2) if incluir_iva27 and i % 7 == 0 else 0.00
    iva_105 = round(gravado * 0.105, 2) if i % 5 == 0 else 0.00
    total = round(gravado + no_gravado + iva_21 + iva_27 + iva_105, 2)

    return [fecha, tipo, pdv, nro, p[0], p[1], gravado, no_gravado, iva_21, iva_27, iva_105, total]


random.seed(42)
filas_ejemplo = [generar_fila(i) for i in range(15)]

# ── 1. Plantilla vacía Contabilidad ──────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "Contabilidad"
ws.sheet_view.showGridLines = False
aplicar_encabezado(ws, "PLANTILLA — Contabilidad", "Completá esta plantilla con los datos exportados desde tu sistema de gestión (ej. Contabilium)")
aplicar_fila_encabezado(ws)

# Fila de ejemplo en azul claro como guía

ejemplo = ["01/06/2025", "FAC A", "0001", "00000001", "20-12345678-9", "PROVEEDOR EJEMPLO SA", 1000.00, 0.00, 210.00, 0.00, 0.00, 1210.00]
escribir_fila_datos(ws, 4, ejemplo, color_bg=COLOR_EJEMPLO)
ws.cell(4, 1).font = Font(name="Arial", size=10, italic=True, color="1A73E8")

ws.freeze_panes = "A4"
wb.save("/home/claude/cruce_iva/ejemplos/plantilla_contabilidad.xlsx")

# ── 2. Plantilla vacía ARCA ───────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "ARCA"
ws.sheet_view.showGridLines = False
aplicar_encabezado(ws, "PLANTILLA — ARCA", "Copiá las columnas necesarias desde el reporte de IVA Compras descargado de ARCA")
aplicar_fila_encabezado(ws)
escribir_fila_datos(ws, 4, ejemplo, color_bg=COLOR_EJEMPLO)
ws.cell(4, 1).font = Font(name="Arial", size=10, italic=True, color="1A73E8")
ws.freeze_panes = "A4"
wb.save("/home/claude/cruce_iva/ejemplos/plantilla_arca.xlsx")

# ── 3. Ejemplo Contabilidad con datos ─────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "Contabilidad"
ws.sheet_view.showGridLines = False
aplicar_encabezado(ws, "Contabilidad — Junio 2025", "Archivo de ejemplo con datos ficticios")
aplicar_fila_encabezado(ws)

for i, fila in enumerate(filas_ejemplo, 4):
    color = "F8F9FA" if i % 2 == 0 else "FFFFFF"
    escribir_fila_datos(ws, i, fila, color_bg=color)

ws.freeze_panes = "A4"
wb.save("/home/claude/cruce_iva/ejemplos/contabilidad_ejemplo.xlsx")

# ── 4. Ejemplo ARCA con datos (con diferencias intencionales) ─────────────

wb = Workbook()
ws = wb.active
ws.title = "ARCA"
ws.sheet_view.showGridLines = False
aplicar_encabezado(ws, "ARCA IVA Compras — Junio 2025", "Archivo de ejemplo con datos ficticios")
aplicar_fila_encabezado(ws)

random.seed(42)
filas_arca = [generar_fila(i) for i in range(15)]

# Introducir diferencias intencionales para el ejemplo

# Fila 2: diferencia en importe_no_gravado

filas_arca[1][7] = 0.00
filas_arca[1][11] = round(filas_arca[1][6] + filas_arca[1][8], 2)

# Fila 5: solo en ARCA (cambiar cuit para que no matchee)

filas_arca[4][4] = "20-99999999-9"

# Fila 8: diferencia en iva_21

filas_arca[7][8] = round(filas_arca[7][8] + 50, 2)
filas_arca[7][11] = round(filas_arca[7][11] + 50, 2)

# Quitar fila 10 de ARCA (solo en contabilidad)

filas_arca.pop(9)

for i, fila in enumerate(filas_arca, 4):
    color = "F8F9FA" if i % 2 == 0 else "FFFFFF"
    escribir_fila_datos(ws, i, fila, color_bg=color)

ws.freeze_panes = "A4"
wb.save("/home/claude/cruce_iva/ejemplos/arca_ejemplo.xlsx")

print("✅ Plantillas y ejemplos generados correctamente.")
