"""
Módulo de exportación de resultados del cruce a un archivo Excel formateado.
Genera una hoja por estado y una hoja de resumen general.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from cruce import COLUMNAS_IMPORTES, COLUMNAS_CRUCE


# ── Paleta de colores ──────────────────────────────────────────────────────
# Colores generales
COLOR_HEADER      = "1A1A2E"
COLOR_SUBHEADER   = "3C3C5A"
# Colores base
COLOR_ALT         = "F8F9FA"
COLOR_WHITE       = "FFFFFF"
# Colores de coincidencia
COLOR_MATCH_BG    = "E6F4EA"
COLOR_MATCH_FT    = "1E8C45"
COLOR_MATCH_HDR   = "1E6B36"
# Colores de diferencias
COLOR_DIFF_BG     = "FFF8E1"
COLOR_DIFF_FT     = "B06000"
COLOR_DIFF_CELDA  = "FFE082"
# Colores Solo Conta
COLOR_SOLO_C_BG   = "FCE8E6"
COLOR_SOLO_C_FT   = "C5221F"
COLOR_SOLO_C_HDR  = "8B1A1A"
# Colores solo ARCA
COLOR_SOLO_A_BG   = "E8F0FE"
COLOR_SOLO_A_FT   = "1A73E8"
COLOR_SOLO_A_HDR  = "1558B0"


_thin = Side(style="thin", color="CCCCCC")
BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

COLUMNAS_SIMPLES = COLUMNAS_CRUCE + ["nombre_empresa"] + COLUMNAS_IMPORTES
COLUMNAS_DOBLES = (
    COLUMNAS_CRUCE
    + ["nombre_empresa"]
    + [f"{c}_cont" for c in COLUMNAS_IMPORTES]
    + [f"{c}_arca" for c in COLUMNAS_IMPORTES]
    + ["dif_total"]
)

ANCHOS_SIMPLES = [14, 18, 14, 16, 20, 28, 14, 16, 10, 10, 12, 16]
ANCHOS_DOBLES  = [14, 18, 14, 16, 20, 28] + [13] * (len(COLUMNAS_IMPORTES) * 2) + [13]


def crear_relleno(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def crear_fuente(color: str = "212529", bold: bool = False, size: int = 10) -> Font:
    return Font(name="Arial", color=color, bold=bold, size=size)


def _escribir_titulo(ws, titulo: str, subtitulo: str, ncols: int) -> None:

    """Escribe el bloque de título en las primeras dos filas de la hoja."""

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = crear_relleno(COLOR_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws["A2"] = subtitulo
    ws["A2"].font = Font(name="Arial", color="888888", size=9, italic=True)
    ws["A2"].fill = crear_relleno(COLOR_HEADER)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16


def _escribir_fila_encabezado(ws, fila: int, columnas: list, color_hdr: str = COLOR_SUBHEADER) -> None:

    """Escribe una fila de encabezados de columna con formato."""

    for col, nombre in enumerate(columnas, 1):
        c = ws.cell(row=fila, column=col, value=nombre)
        c.font = crear_fuente("FFFFFF", bold=True)
        c.fill = crear_relleno(color_hdr)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDE
    ws.row_dimensions[fila].height = 20


def _escribir_fila_datos(ws, fila: int, valores: list, color_bg: str, color_ft: str = "333333") -> None:

    """Escribe una fila de datos con fondo alternado."""

    numericas = set(COLUMNAS_IMPORTES) | {f"{c}_cont" for c in COLUMNAS_IMPORTES} | {f"{c}_arca" for c in COLUMNAS_IMPORTES} | {"dif_total"}
    for col, (nombre_col, valor) in enumerate(valores, 1):
        c = ws.cell(row=fila, column=col, value=valor)
        c.font = crear_fuente(color_ft)
        c.fill = crear_relleno(color_bg)
        c.border = BORDE
        c.alignment = Alignment(
            horizontal="right" if nombre_col in numericas else "left",
            vertical="center"
        )
        if nombre_col in numericas and isinstance(valor, (int, float)):
            c.number_format = "#,##0.00"
    ws.row_dimensions[fila].height = 16


def _set_anchos(ws, anchos: list) -> None:

    """Ajusta el ancho de cada columna."""

    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _fila_a_lista(clave, datos: dict | None, columnas: list) -> list:

    """Convierte una clave de índice y un dict de datos en lista ordenada."""

    fila = dict(zip(COLUMNAS_CRUCE, clave if isinstance(clave, tuple) else (clave,)))
    if datos:
        fila.update(datos)
    return [(col, fila.get(col, None)) for col in columnas]


def _escribir_total(ws, fila: int, n_col_fijas: int, n_cols: int, color_bg: str, color_ft: str) -> None:

    """Escribe una fila de totales con fórmulas SUM al final de la tabla."""

    for columna in range(1, n_col_fijas + 1):
        celda = ws.cell(row=fila, column=columna)
        celda.fill = crear_relleno(color_bg)
        celda.border = BORDE
    celda_total = ws.cell(row=fila, column=n_col_fijas, value="TOTAL")
    celda_total.font = crear_fuente("555555", bold=True)
    celda_total.alignment = Alignment(horizontal="right")

    for columna in range(n_col_fijas + 1, n_cols + 1):
        letra_columna = get_column_letter(columna)

        celda = ws.cell(
            row=fila,
            column=columna,
            value=f"=SUM({letra_columna}4:{letra_columna}{fila - 1})",
        )
        

        celda.font = crear_fuente(color_ft, bold=True)
        celda.fill = crear_relleno(color_bg)
        celda.border = BORDE
        celda.alignment = Alignment(horizontal="right", vertical="center")
        celda.number_format = "#,##0.00"

    ws.row_dimensions[fila].height = 18


# ── Hojas individuales ─────────────────────────────────────────────────────

def _hoja_simples(wb: Workbook, titulo: str, subtitulo: str, filas: list,
                    color_bg: str, color_ft: str, color_hdr: str) -> None:
    """Crea una hoja con formato simple (una sola fuente de datos)."""
    ws = wb.create_sheet(titulo)
    ws.sheet_view.showGridLines = False
    _set_anchos(ws, ANCHOS_SIMPLES)
    _escribir_titulo(ws, titulo, subtitulo, len(COLUMNAS_SIMPLES))
    _escribir_fila_encabezado(ws, 3, COLUMNAS_SIMPLES, color_hdr)

    for i, r in enumerate(filas, 4):
        datos = r["contabilidad"] or r["arca"]
        clave = r["clave"]
        valores = _fila_a_lista(clave, datos, COLUMNAS_SIMPLES)
        alt = COLOR_ALT if i % 2 == 0 else COLOR_WHITE
        _escribir_fila_datos(ws, i, valores, alt, color_ft)

    if filas:
        _escribir_total(ws, len(filas) + 4, len(COLUMNAS_CRUCE) + 1,
                        len(COLUMNAS_SIMPLES), color_bg, color_ft)
    ws.freeze_panes = "A4"


def _hoja_diferencias(wb: Workbook, filas: list) -> None:
    """Crea la hoja de diferencias con columnas comparativas lado a lado."""
    ws = wb.create_sheet("⚠️ Diferencias")
    ws.sheet_view.showGridLines = False
    _set_anchos(ws, ANCHOS_DOBLES)
    n_fijas = len(COLUMNAS_CRUCE) + 1  # cruce + nombre_empresa

    titulo = f"⚠️ Diferencias ({len(filas)} comprobantes)"
    subtitulo = "Celdas en amarillo indican campos con diferencia entre Contabilidad y ARCA"
    _escribir_titulo(ws, titulo, subtitulo, len(COLUMNAS_DOBLES))

    # Sub-encabezados agrupados
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_fijas)
    ws["A3"] = "Identificación del comprobante"
    ws["A3"].font = crear_fuente("FFFFFF", bold=True)
    ws["A3"].fill = crear_relleno(COLOR_SUBHEADER)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    col_cont_ini = n_fijas + 1
    col_cont_fin = n_fijas + len(COLUMNAS_IMPORTES)
    ws.merge_cells(start_row=3, start_column=col_cont_ini, end_row=3, end_column=col_cont_fin)
    ws.cell(3, col_cont_ini, "📒 CONTABILIDAD").font = crear_fuente("FFFFFF", bold=True)
    ws.cell(3, col_cont_ini).fill = crear_relleno(COLOR_MATCH_HDR)
    ws.cell(3, col_cont_ini).alignment = Alignment(horizontal="center", vertical="center")

    col_arca_ini = col_cont_fin + 1
    col_arca_fin = col_cont_fin + len(COLUMNAS_IMPORTES)
    ws.merge_cells(start_row=3, start_column=col_arca_ini, end_row=3, end_column=col_arca_fin)
    ws.cell(3, col_arca_ini, "🏛 ARCA").font = crear_fuente("FFFFFF", bold=True)
    ws.cell(3, col_arca_ini).fill = crear_relleno(COLOR_SOLO_A_HDR)
    ws.cell(3, col_arca_ini).alignment = Alignment(horizontal="center", vertical="center")

    col_dif = col_arca_fin + 1
    ws.cell(3, col_dif, "Dif. Total").font = crear_fuente("FFFFFF", bold=True)
    ws.cell(3, col_dif).fill = crear_relleno("8A6000")
    ws.cell(3, col_dif).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(3, col_dif).border = BORDE

    _escribir_fila_encabezado(ws, 4, COLUMNAS_DOBLES, COLOR_SUBHEADER)

    for i, r in enumerate(filas, 5):
        clave = r["clave"]
        cont = r["contabilidad"] or {}
        arca = r["arca"] or {}
        alt = COLOR_ALT if i % 2 == 0 else COLOR_WHITE

        # Columnas de identificación
        fila_base = dict(zip(COLUMNAS_CRUCE, clave if isinstance(clave, tuple) else (clave,)))
        fila_base["nombre_empresa"] = (cont or arca).get("nombre_empresa", "")

        for col, nombre in enumerate(COLUMNAS_CRUCE + ["nombre_empresa"], 1):
            c = ws.cell(row=i, column=col, value=fila_base.get(nombre, ""))
            c.font = crear_fuente()
            c.fill = crear_relleno(alt)
            c.border = BORDE
            c.alignment = Alignment(horizontal="left", vertical="center")

        # Columnas contabilidad
        for j, col_imp in enumerate(COLUMNAS_IMPORTES):
            col = col_cont_ini + j
            val = cont.get(col_imp, 0)
            c = ws.cell(row=i, column=col, value=val)
            c.font = crear_fuente()
            c.fill = crear_relleno(alt)
            c.border = BORDE
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0.00"

        # Columnas ARCA (resaltar diferencias)
        for j, col_imp in enumerate(COLUMNAS_IMPORTES):
            col = col_arca_ini + j
            val_arca = arca.get(col_imp, 0)
            val_cont = cont.get(col_imp, 0)
            es_dif = col_imp in r["diferencias"]
            c = ws.cell(row=i, column=col, value=val_arca)
            c.font = crear_fuente(bold=es_dif)
            c.fill = crear_relleno(COLOR_DIFF_CELDA if es_dif else alt)
            c.border = BORDE
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0.00"

        # Diferencia de total
        letra_cont = get_column_letter(col_cont_ini + COLUMNAS_IMPORTES.index("importe_total"))
        letra_arca = get_column_letter(col_arca_ini + COLUMNAS_IMPORTES.index("importe_total"))
        c = ws.cell(row=i, column=col_dif, value=f"={letra_cont}{i}-{letra_arca}{i}")
        c.font = crear_fuente(COLOR_DIFF_FT, bold=True)
        c.fill = crear_relleno(COLOR_DIFF_CELDA)
        c.border = BORDE
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = "#,##0.00"

        ws.row_dimensions[i].height = 16

    ws.freeze_panes = "A5"


def _hoja_resumen(wb: Workbook, resumen: dict) -> None:
    """Crea la hoja de resumen con los conteos por estado."""
    ws = wb.create_sheet("📊 Resumen", 0)
    ws.sheet_view.showGridLines = False
    for col, ancho in zip("ABCDE", [28, 18, 22, 22, 22]):
        ws.column_dimensions[col].width = ancho

    _escribir_titulo(ws, "CRUCE IVA COMPRAS", "Resumen de resultados por estado", 5)

    encabezados = ["Estado", "Comprobantes", "% del total", "Total Contabilidad", "Total ARCA"]
    _escribir_fila_encabezado(ws, 3, encabezados, COLOR_SUBHEADER)

    estados = [
        ("✅ Match",              "match",             COLOR_MATCH_BG,   COLOR_MATCH_FT),
        ("⚠️ Diferencia",         "diferencia",        COLOR_DIFF_BG,    COLOR_DIFF_FT),
        ("❌ Solo Contabilidad",  "solo_contabilidad", COLOR_SOLO_C_BG,  COLOR_SOLO_C_FT),
        ("🔵 Solo ARCA",          "solo_arca",         COLOR_SOLO_A_BG,  COLOR_SOLO_A_FT),
    ]

    for fila, (label, key, bg, ft) in enumerate(estados, 4):
        cant = resumen.get(key, 0)
        total = resumen.get("total", 1)
        pct = cant / total if total else 0

        ws.cell(fila, 1, label).font = crear_fuente(ft, bold=True)
        ws.cell(fila, 1).fill = crear_relleno(bg); ws.cell(fila, 1).border = BORDE
        ws.cell(fila, 1).alignment = Alignment(horizontal="left", vertical="center")

        ws.cell(fila, 2, cant).font = crear_fuente(ft, bold=True)
        ws.cell(fila, 2).fill = crear_relleno(bg); ws.cell(fila, 2).border = BORDE
        ws.cell(fila, 2).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(fila, 3, pct).font = crear_fuente(ft)
        ws.cell(fila, 3).fill = crear_relleno(bg); ws.cell(fila, 3).border = BORDE
        ws.cell(fila, 3).number_format = "0.0%"
        ws.cell(fila, 3).alignment = Alignment(horizontal="center", vertical="center")

        for col in [4, 5]:
            ws.cell(fila, col, "—").font = crear_fuente("AAAAAA")
            ws.cell(fila, col).fill = crear_relleno(bg); ws.cell(fila, col).border = BORDE
            ws.cell(fila, col).alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[fila].height = 18

    # Fila total
    ws.cell(8, 1, "TOTAL GENERAL").font = crear_fuente("FFFFFF", bold=True)
    ws.cell(8, 1).fill = crear_relleno(COLOR_SUBHEADER); ws.cell(8, 1).border = BORDE
    ws.cell(8, 2, resumen.get("total", 0)).font = crear_fuente("FFFFFF", bold=True)
    ws.cell(8, 2).fill = crear_relleno(COLOR_SUBHEADER); ws.cell(8, 2).border = BORDE
    ws.cell(8, 2).alignment = Alignment(horizontal="center", vertical="center")
    for col in [3, 4, 5]:
        ws.cell(8, col).fill = crear_relleno(COLOR_SUBHEADER); ws.cell(8, col).border = BORDE
    ws.row_dimensions[8].height = 20


# ── Función principal de exportación ──────────────────────────────────────

def exportar_excel(resultados: list[dict], resumen_data: dict) -> bytes:
    """
    Genera el archivo Excel de resultados del cruce y lo devuelve como bytes.

    Args:
        resultados: Lista de resultados devuelta por cruce.cruzar().
        resumen_data: Diccionario de conteos devuelto por cruce.resumen().

    Returns:
        Contenido del archivo Excel como bytes, listo para descargar.
    """
    wb = Workbook()
    wb.remove(wb.active)

    matches        = [r for r in resultados if r["estado"] == "match"]
    diferencias    = [r for r in resultados if r["estado"] == "diferencia"]
    solo_cont      = [r for r in resultados if r["estado"] == "solo_contabilidad"]
    solo_arca      = [r for r in resultados if r["estado"] == "solo_arca"]

    _hoja_resumen(wb, resumen_data)

    _hoja_simples(
        wb,
        titulo=f"✅ Match ({len(matches)})",
        subtitulo="Comprobantes que cruzaron correctamente en ambos archivos",
        filas=matches,
        color_bg=COLOR_MATCH_BG, color_ft=COLOR_MATCH_FT, color_hdr=COLOR_MATCH_HDR,
    )

    if diferencias:
        _hoja_diferencias(wb, diferencias)

    if solo_cont:
        _hoja_simples(
            wb,
            titulo=f"❌ Solo Contabilidad ({len(solo_cont)})",
            subtitulo="Comprobantes presentes en Contabilidad pero no encontrados en ARCA",
            filas=solo_cont,
            color_bg=COLOR_SOLO_C_BG, color_ft=COLOR_SOLO_C_FT, color_hdr=COLOR_SOLO_C_HDR,
        )

    if solo_arca:
        _hoja_simples(
            wb,
            titulo=f"🔵 Solo ARCA ({len(solo_arca)})",
            subtitulo="Comprobantes presentes en ARCA pero no registrados en Contabilidad",
            filas=solo_arca,
            color_bg=COLOR_SOLO_A_BG, color_ft=COLOR_SOLO_A_FT, color_hdr=COLOR_SOLO_A_HDR,
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()